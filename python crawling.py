
──────────────────── 목차 ────────────────────

------------- 라이브러리 -------------
------------- 업데이트 내용 -------------
------------- 크롤링 설명 -------------



############# 0.크롤링 여러 기능 #############
############# 1.yes24 크롤링 #############
############# 2.네이버쇼핑 크롤링 #############

──────────────────────────────────────────────



------------- 라이브러리 -------------

# 기본
import requests

import warnings
warnings.filterwarnings('ignore')

# 셀레늄
pip install chromedriver_autoinstaller
import chromedriver_autoinstaller

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions


# 뷰티풀솝
from bs4 import BeautifulSoup

# 그 외 동작기능
import time
from urllib.request import urlretrieve

# 엑셀
import openpyxl

# 데이터분석, 데이터전처리
import pandas as pd

# 데이터 시각화
import matplotlib.pyplot as plt
from matplotlib import rc
plt.rc('font', family='malgun gothic')
plt.rc('axes', unicode_minus=False)
import seaborn as sns

# 저장
import datetime

# DB?
from const import DB_PATH






------------- 업데이트 내용 -------------


셀레니움3에서는 아래의 방법으로 html 요소를 찾았다면
driver.find_element_by_class_name("")
driver.find_element_by_id("")
driver.find_element_by_css_selector("")
driver.find_element_by_name("")
driver.find_element_by_tag_name("")
driver.find_element_by_xpath("")
driver.find_element_by_link_text("")
driver.find_element_by_partial_link_text("")
(복수형 driver.find_elements_by~~)

셀레니움4에서는
driver.find_element(By.CLASS_NAME, "")
driver.find_element(By.ID, "")
driver.find_element(By.CSS_SELECTOR, "")
driver.find_element(By.NAME, "")
driver.find_element(By.TAG_NAME, "")
driver.find_element(By.XPATH, "")
driver.find_element(By.LINK_TEXT, "")
driver.find_element(By.PARTIAL_LINK_TEXT, "")
(복수형 driver.find_elements(By.~~, "")
이렇게 바뀌었습니다.




------------- 크롤링 설명 -------------

 - CSS_SELECTOR는 다 되죠.
★★★★★ ID면 앞에 (샵), CLASS_NAME이면 앞에 .을 입력해주면 됩니다.
          css_selector로 사실 거의 다 작성함.

 - 내부에 있는 title이나 등등 모두 활용이 가능.
driver.find_element(By.CSS_SELECTOR, "[title='검색어 입력']").send_keys("뉴진스")
   → 검색어 입력이란 타이틀을 가진 것을 불러와라. 라는 것.
driver.find_element(By.CSS_SELECTOR, "[placeholder='검색어를 입력해 주세요.']").send_keys("에스파")

 - XPATH
   → xpath는 최후의 방법으로 생각하세요. 가능하면 앞에 것들도 해결하는게 훨씬 좋습니다
driver.find_element(By.XPATH, '//*[@id="query"]').send_keys("에스파")


 - LINK_TEXT
driver.find_element(By.LINK_TEXT, "쇼핑LIVE").click()
  → 쇼핑LIVE라는 글자가 있는 것을 가져옴

  ★★ 하나가 더 있음. 파셜 링크 텍스트로, 글자의 일부가 해당하는 것이 있어도 찾아내는 기능이에요.
driver.find_element(By.PARTIAL_LINK_TEXT, "핑LI").click()

 - TAG_NAME
# 태그가 많다보니 단독으로 있을 때 사용함. 이용도로는 자주 사용안함
# 혹은 find_elements로 여러 개 있을 때 사용하곤 함
driver.find_element(By.TAG_NAME, "")


    # 예시
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time

options = Options()
options.add_experimental_option("detach", True)
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=options)

url = "https://naver.com"

driver.get(url)
time.sleep(1)

    # CLASS_NAME
driver.find_element(By.CLASS_NAME, "input_text").send_keys("블랙핑크")
time.sleep(1)
    # ID
driver.find_element(By.ID, "query").send_keys("뉴진스")
time.sleep(1)
    # NAME
driver.find_element(By.NAME, "query").send_keys("트와이스")
time.sleep(1)
    # CSS_SELECTOR
driver.find_element(By.CSS_SELECTOR, "#query").send_keys("에스파")
time.sleep(1)
    # XPATH
driver.find_element(By.XPATH, '//*[@name="query"]').send_keys("에스파")
time.sleep(1)
    # LINK_TEXT
driver.find_element(By.LINK_TEXT, "쇼핑LIVE").click()
    # PARTIAL_LINK_TEXT
driver.find_element(By.PARTIAL_LINK_TEXT, "핑LI").click()
    # CSS_SELECTOR
navs = driver.find_elements(By.CSS_SELECTOR, ".nav")

for nav in navs:
    print(nav.get_attribute("outerHTML"))
    print()



 - WebDriverWait
WebDriverWait(driver, 10) : WebDriverWait을 통해 driver를 "최대" 10초동안 기다린다.
(이때 10초를 넘기면 NoSuchElementException, ElementNotVisibleException과 같은 에러가 발생한다.)
이런 에러가 뜬다면, 한번 Waits을 해보자)

.until(EC.presence_of_element_located((By.ID, "myDynamicElement")) : 언제까지? -> ID가 "myDynamicElement"인 엘리먼트가 나올 때까지
(이때 해당 엘리먼트가 나오면, EC에서 True를 리턴한다.)
EC에서 presence_of_element_located 말고도 다양한 메소드를 제공한다.(ex. element_to_be_clickable() 등등..)

즉, ID가 myDynamicElement인 element가 나올 때까지 최대 10초를 기다린다.
이때, ID 이외에도 XPATH, CLASS_NAME, LINK_TEXT 등등을 이용해 Explicit Waits을 구현 할 수 있다.
그냥 저 틀을 알아두면 편할 것 같다.

 - maximize_window
browser.maximize_window() # 창 최대화







############# 0.크롤링 여러 기능 #############

select_element = driver.find_element(By.NAME, 'selectomatic')
select = Select(select_element)

# 보이는 텍스트 기반 옵션 선택
select.select_by_visible_text('Four')

# 값 속성에 따라 옵션 선택
select.select_by_value('two')

# 목록에서 해당 색인에 따라 옵션 선택
select.select_by_index(3)





############# 1.yes24 크롤링 #############

## 과제

- yes24에서 python을 검색하고 200권의 책의 정보를 수집하는 코드를 작성하고, csv파일과 xlsx파일로 저장하시오
  (다른 책을 검색해도 됨 / 과제 예시는 python을 검색한 코드로 나갈 것)
- 책사진, 제목, 저자, 가격, 판매지수, 회원리뷰수, 별점 점수, 사이크 링크를 모두 포함해야함
- 책사진은 폴더를 따로 만들어 저장한 후 캡처본으로 제출
- 패키지 이용에는 관여하지 않음
- 제출 목록(4개) : 4차시과제_학번.ipynb파일, 4차시과제_학번.csv파일, 4차시과제_학번.excel파일, 4차시과제_학번.사진저장 폴더의 캡처본

#### Libraries

import requests
from bs4 import BeautifulSoup

import time
from urllib.request import urlretrieve

from selenium import webdriver
from selenium.webdriver.common.keys import Keys

import openpyxl


# 1. Requests와 BeautifulSoup를 사용한 크롤링

# 검색, 페이지
keyword = 'python'; page =1
# 스크랩한 책의 갯수
total_books = 1
# 제목 / 저자 /가격 / 판매지수 / 회원리뷰수 / 별점 점수 / 링크 리스트
titles = []; writers = []; prices = []; sales_indexs = [];
views = []; rating_scores = []; links = [];


while total_books < 200:
    req = requests.get(f'http://www.yes24.com/searchcorner/Search?keywordAd=&keyword=&domain=BOOK&qdomain=%b1%b9%b3%bb%b5%b5%bc%ad&query={keyword}&PageNumber={page}',headers={'User_Agent':'Mozilla/5.0'})
    print(f'page: {page}, req status: {req.status_code}')

    # html 구조로 인식
    soup = BeautifulSoup(req.text, 'html.parser')

    # ----- image 저장
    images = soup.select('td.goods_img')
    for image in images:
        image_src = image.select_one('a img').attrs['src']
        urlretrieve(image_src, f'./product_pictures/beautifulSoup/python/{total_books}.jpg'); total_books +=1
    # ----- 제목 / 저자 /가격 / 판매지수 / 회원리뷰수 / 별점 점수 / 링크 저장
    books = soup.select('td.goods_infogrp')
    for book in books:
        title = book.select_one('p a').text
        titles.append(title.replace(",", "-"))
        if ' 저' in book.select_one('div.goods_info').text:
            writer = book.select_one('div.goods_info').text.split(' 저')[0].strip('\n')
            writers.append(writer.replace(",", "-"))
        elif ' 공저' in book.select_one('div.goods_info').text:
            writer = book.select_one('div.goods_info').text.split(' 공저')[0].strip('\n')
            writers.append(writer.replace(",", "-"))
        elif ' 편저' in book.select_one('div.goods_info').text:
            writer = book.select_one('div.goods_info').text.split(' 편저')[0].strip('\r\n                ')
            writers.append(writer.replace(",", "-"))
        price = book.select_one('div.goods_price em').text
        prices.append(price.replace(",", ""))
        try:
            sales_index = book.select_one('span.gd_reviewCount').text.split()[1]
            sales_indexs.append(sales_index.replace(",", ""))
        except:
            sales_indexs.append('판매지수없음')
        try:
            view = book.select_one('div.goods_rating span a em').text
            views.append(view.replace(",", ""))
        except:
            views.append('0')
        try:
            rating_score = book.select_one('span.gd_rating em').text
            rating_scores.append(rating_score)
        except:
            rating_scores.append('별점미등록')
        href = book.select_one('p a').attrs['href']
        link = 'http://www.yes24.com' + href
        links.append(link)
    page += 1

print('Crawling Finished !')

# csv 파일로 저장
f = open('yes24_csv_python.csv', 'w')
f.write('제목, 저자, 가격, 판매지수, 회원리뷰수, 별점 점수, 링크 리스트\n')
for i in range(total_books-1):
    f.write(f'{titles[i]}, {writers[i]}, {prices[i]}, {sales_indexs[i]}, {views[i]}, {rating_scores[i]}, {links[i]}\n')
f.close()

# excel 파일로 저장
wb = openpyxl.Workbook()
sheet = wb.active
sheet.append('제목, 저자, 가격, 판매지수, 회원리뷰수, 별점 점수, 링크 리스트'.split(','))
for i in range(total_books-1):
    sheet.append(f'{titles[i]}, {writers[i]}, {prices[i]}, {sales_indexs[i]}, {views[i]}, {rating_scores[i]}, {links[i]}'.split(','))
wb.save('yes24_excel_python.xlsx')

# 2. Selenium을 사용한 크롤링

# 드라이버 불러오기(크롬으로 빈 페이지가 불러와짐)
dr = webdriver.Chrome('./chromedriver_win32/chromedriver')

# 원하는 페이지 get메소드로 불러오기
dr.get('http://www.yes24.com/')

# 검색입력
search_bar = dr.find_element_by_css_selector('span.schIpt label input')
search_bar.send_keys('python')

# 클릭하기
search = dr.find_element_by_css_selector('span.schBtn button')
search.click()

# 스크랩한 책의 갯수, 페이지
total_books = 1; page = 1
# 제목 / 저자 /가격 / 판매지수 / 회원리뷰수 / 별점 점수 / 링크 리스트
titles = []; writers = []; prices = []; sales_indexs = [];
views = []; rating_scores = []; links = [];

while total_books < 200 :
    # ----- 이미지 저장
    images = dr.find_elements_by_css_selector('td.goods_img')[:20]
    for image in images:
        image_src = image.find_element_by_css_selector('a img').get_attribute('src')
        urlretrieve(image_src, f'./product_pictures/selenium/python/{total_books}.jpg')
        total_books +=1
    # ----- 제목 / 저자 /가격 / 판매지수 / 회원리뷰수 / 별점 점수 / 링크 저장
    books = dr.find_elements_by_css_selector('td.goods_infogrp')[:20]
    for book in books:
        title = book.find_element_by_css_selector('p a').text
        titles.append(title.replace(",","-"))
        writer_infos = book.find_element_by_css_selector('div.goods_info').text
        if ' 저' in writer_infos:
            writer = writer_infos.split(' 저')[0]
            writers.append(writer.replace(",","-"))
        elif ' 공저' in writer_infos:
            writer = writer_infos.split(' 공저')[0]
            writers.append(writer.replace(",","-"))
        elif ' 편저' in writer_infos:
            writer = writer_infos.split(' 편저')[0]
            writers.append(writer)
        elif ' 역' in writer_infos:
            writer = writer_infos.split(' 역')[0]
            writers.append(writer)
        else:
            writer = writer_infos.split(' |')[0]
            writers.append(writer.replace(",","-"))
        price = book.find_element_by_css_selector('div.goods_price em').text.replace(",", "")
        prices.append(price.replace(", ", ""))
        try:
            sales_index = book.find_element_by_css_selector('div.goods_rating > span').text.split()[1]
            sales_indexs.append(sales_index.replace(",",""))
        except:
            sales_indexs.append('판매지수없음')
        try:
            view = book.find_element_by_css_selector('div.goods_rating span a em').text.replace(", ", "")
            views.append(view.replace(",",""))
        except:
            views.append('0')
        try:
            rating_score = book.find_element_by_css_selector('span.gd_rating em').text
            rating_scores.append(rating_score)
        except:
            rating_scores.append('별점미등록')
        link = book.find_element_by_css_selector('p a').get_attribute('href')
        links.append(link)
    print(f'page: {page}')
    page += 1
    try:
        next_page = dr.find_element_by_css_selector(f'a.n:nth-of-type({page+1})')
        next_page.click()
        time.sleep(1.5)
    except:
        break

print('Crawling Finished !')

# csv 파일로 저장
f = open('yes24_csv_python_selenium.csv', 'w')
f.write('제목, 저자, 가격, 판매지수, 회원리뷰수, 별점 점수, 링크 리스트\n')
for i in range(total_books-1):
    f.write(f'{titles[i]}, {writers[i]}, {prices[i]}, {sales_indexs[i]}, {views[i]}, {rating_scores[i]}, {links[i]}\n')
f.close()

# excel 파일로 저장
wb = openpyxl.Workbook()
sheet = wb.active
sheet.append('제목, 저자, 가격, 판매지수, 회원리뷰수, 별점 점수, 링크 리스트'.split(','))
for i in range(total_books-1):
    sheet.append(f'{titles[i]}, {writers[i]}, {prices[i]}, {sales_indexs[i]}, {views[i]}, {rating_scores[i]}, {links[i]}'.split(','))
wb.save('yes24_excel_python_selenium.xlsx')

**Fin───────────────────────────────────────────────────────────────────────────────────────────────**

############# 2.네이버쇼핑 크롤링 #############


## 과제
- 네이버 쇼핑에서 원하는 상품을 검색하고, 500개 이상의 데이터를
수집하는 코드를 작성하고, csv파일과 xlsx파일로 저장하시오
- 제품사진, 제품명, 가격, 리뷰수, 구매건수, 찜한수, 사이트 링크를 모두 포함해야함
- 제품사진은 폴더를 따로 만들어 저장한 후 캡처본으로 제출
- 패키지 이용에는 관여하지 않음
- 제일 처음 받는 웹페이지의 주소는 네이버 쇼핑의 첫 화면이어야 함
   (requests&BeautifulSoup의 경우에는 keyword를 받아서 검색한 창이 처음에 떠도 됨)
- 제출 목록 : ipynb파일, csv파일, excel파일, 사진저장 폴더의 캡처본

#### Libraries

import requests
from bs4 import BeautifulSoup

import time
from urllib.request import urlretrieve

from selenium import webdriver
from selenium.webdriver.common.keys import Keys # 스크롤 내리기

import openpyxl

# Selenium을 사용한 크롤링

# 드라이버 불러오기(크롬으로 빈 페이지가 불러와짐)
dr = webdriver.Chrome('./chromedriver_win32/chromedriver')

# 원하는 페이지 get메소드로 불러오기
dr.get('https://shopping.naver.com/')

# 검색입력
search_bar = dr.find_element_by_css_selector('#autocompleteWrapper > input.co_srh_input._input.N\=a\:SNB\.search')
search_bar.send_keys('나이키 데이브레이크')

# 클릭하기
search = dr.find_element_by_css_selector('#autocompleteWrapper > a.co_srh_btn._search.N\=a\:SNB\.search')
search.click()

# ----- 크롤링할 전체 제품 수
products_num = 0
# ----- 제품명 / 가격 / 리뷰수 / 구매건수 / 찜한수 / 제품링크
name_list = []
price_list = []
views_list = []
buys_list = []
jjim_list = []
link_list = []
# ----- 페이지 body
body = dr.find_element_by_css_selector('body')

# ----- 제품이 100개가 넘어가면 스탑
while products_num < 100:
    # 페이지 넘어간 후 첫 화면 이미지 로드
    time.sleep(1.2)

    # 스크롤 12번 내림: 이미지 로드하는 과정
    for scroll in range(12):
        body.send_keys(Keys.PAGE_DOWN)

        time.sleep(0.2)

    # 페이지에 나온 제품들 받음
    products = dr.find_elements_by_css_selector('#__next > div > div.style_container__1YjHN > div.style_inner__18zZX > div.style_content_wrap__1PzEo > div.style_content__2T20F > ul > div > div > li')
    print(len(products))
    for product in products:
        # ----- 이미지 저장
        image = product.find_element_by_css_selector('a.thumbnail_thumb__3Agq6 img').get_attribute('src')
        urlretrieve(image, f'./product_pictures/selenium/nike_DayBreak_{products_num}.jpg'); products_num +=1
        # ----- 제품명 / 가격 / 리뷰수 / 구매건수 / 찜한 수 / 링크 리스트에 저장
        name = product.find_element_by_css_selector('div.basicList_title__3P9Q7 a').text
        name_list.append(name)

        price = product.find_element_by_css_selector('span.price_num__2WUXn').text.replace(',', '')
        price_list.append(price)

        views_buys = product.find_elements_by_css_selector('a.basicList_etc__2uAYO')
        if len(views_buys) == 0:
            views = '0'; buys = '0'
        else:
            for v_b in views_buys:
                if v_b.text.startswith('리뷰'):
                    views = v_b.find_element_by_css_selector('em.basicList_num__1yXM9').text.replace(',','')
                else:
                    buys = v_b.text[4:].replace(',','')
        views_list.append(views); buys_list.append(buys); buys = '0'

        jjim = product.find_element_by_css_selector('span em.basicList_num__1yXM9').text.replace(',','')
        jjim_list.append(jjim)

        link = product.find_element_by_css_selector('div.basicList_title__3P9Q7 a').get_attribute('href')
        link_list.append(link)

    # ----- 다음 페이지로 넘어가기
    next_page = dr.find_element_by_css_selector('a.pagination_next__1ITTf')
    next_page.click()

print('Crawling Finished !')


# csv 파일로 저장
f = open('selenium_test.csv', 'w')
f.write('제품명, 가격, 리뷰, 구매건수, 찜한 수, 사이트링크\n')
for i in range(products_num):
    if name_list[i].startswith('[국내매장판]나이키 데이브레이크 우먼스'): # get \(backslash) in name
        continue
    else:
        f.write(f'{name_list[i]}, {price_list[i]}, {views_list[i]}, {buys_list[i]}, {jjim_list[i]}, {link_list[i]}\n')
f.close()

# excel 파일로 저장
wb = openpyxl.Workbook()
sheet = wb.active
sheet.append('제품명, 가격, 리뷰, 구매건수, 찜한 수, 사이트링크'.split(','))
for i in range(products_num):
    sheet.append(f'{name_list[i]}, {price_list[i]}, {views_list[i]}, {buys_list[i]}, {jjim_list[i]}, {link_list[i]}'.split(','))
wb.save('selenium_test_excel.xlsx')




# 3.파티베이터 코드 일부 #
import datetime
import time
import pandas
import selenium import webdriver
from selenium.webdriver.common.by import By

from const import DB_PATH

def get_moim():
    # 여기야 ========================
    SEARCHING_TEXT = "모임 공지"
    CATEGORY = "카페"
    SCROLL = 5
    # ============================

    browser = webdriver.Chrome()
    browser.get("https://www.naver.com/")
    time.sleep(3)


# 2.cafe_top12
 * 이사람이 검색한건 '고기', '독서', '바베큐', '아이돌', '보드게임', '와인', '위스키' 이런 키워드가 있었음

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By

import time
import pandas
import datetime

# ===================
search_word = "팬카페"
# ===================

driver = webdriver.Chrome()
driver.get('https://section.cafe.naver.com/ca-fe/home/search/combinations?q=')

search = driver.find_element(By.XPATH, '//*[@id="header"]/div[1]/div/div[2]/form/fieldset/div/div/div[2]/input')
search.sesnd_keys(search_word)
search.send_keys(Keys.ENTER)
time.sleep(1)

driver.find_element(By.XPAYH, '//*[@id="mainContaineer"]/div/.div[1]//div[2]/a[2]').click()
time.sleep(1)
